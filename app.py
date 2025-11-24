# -*- coding: utf-8 -*-
"""
Aplicație web pentru procesarea comenzilor de decanturi
OBSID - Platformă de management decanturi parfumuri
"""

import eventlet
eventlet.monkey_patch()

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_socketio import SocketIO, emit
import pandas as pd
import re
from collections import defaultdict
from pathlib import Path
import os
from dotenv import load_dotenv
import database

# Încarcă variabilele din .env dacă există
load_dotenv()

from datetime import datetime
from werkzeug.utils import secure_filename
import io
import logging
import threading
import eventlet
import eventlet.queue as queue
import time
import requests

# Configurare logging
# Update: Added more logging for debugging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Încărcare bază de date produse (SKU mapping)
PRODUCT_DB = {}
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/17FhRBDaknpXgsoTXOkpEWcMf2o55uOjDymlaGiiKUwU/export?format=csv&gid=1884124540"

def normalize_name(text):
    """
    Normalizează numele produsului pentru matching:
    - lowercase
    - elimină 'parfum'
    - elimină caractere non-alfanumerice
    """
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = text.replace('parfum', '')
    text = re.sub(r'[^a-z0-9]', '', text)
    return text

def load_product_db():
    global PRODUCT_DB
    df_db = None
    
    # 1. Încercare încărcare din Google Sheets
    try:
        logger.info(f"🌐 Încercare descărcare bază de date din Google Sheets...")
        response = requests.get(GOOGLE_SHEET_URL, timeout=10)
        response.raise_for_status()
        
        df_db = pd.read_csv(io.BytesIO(response.content))
        logger.info("✅ Baza de date descărcată cu succes din Google Sheets!")
        
        # Salvare cache local (backup)
        try:
            df_db.to_excel('produse.xlsx', index=False)
            logger.info("💾 Cache local actualizat (produse.xlsx)")
        except Exception as e:
            logger.warning(f"⚠ Nu s-a putut salva cache-ul local: {e}")
            
    except Exception as e:
        logger.error(f"❌ Eroare la descărcarea din Google Sheets: {e}")
        logger.info("📂 Încercare încărcare din cache local (produse.xlsx)...")
    
    # 2. Fallback la fișier local dacă descărcarea a eșuat
    if df_db is None:
        try:
            possible_paths = [
                os.path.join(os.path.dirname(__file__), 'produse.xlsx'),
                os.path.join(os.getcwd(), 'produse.xlsx'),
                'produse.xlsx'
            ]
            
            db_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    db_path = path
                    break
            
            if db_path:
                logger.info(f"📂 Încărcare bază de date din local: {db_path}")
                df_db = pd.read_excel(db_path)
            else:
                logger.error("❌ CRITIC: Nu s-a găsit nici baza de date online, nici cea locală!")
                return
        except Exception as e:
            logger.error(f"❌ Eroare la încărcarea locală: {e}")
            return

    # 3. Procesare date
    try:
        count = 0
        PRODUCT_DB = {} # Reset DB
        for _, row in df_db.iterrows():
            nume = str(row['Denumire Produs'])
            sku = str(row['Cod Produs (SKU)']).strip()
            if nume and sku and sku.lower() != 'nan':
                norm_nume = normalize_name(nume)
                PRODUCT_DB[norm_nume] = sku
                count += 1
        logger.info(f"✅ Baza de date produse activă: {count} produse")
    except Exception as e:
        logger.error(f"❌ Eroare la procesarea datelor din DB: {e}")

app = Flask(__name__)

# Încărcăm baza de date la pornire
with app.app_context():
    load_product_db()
    database.init_db()
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['EXPORT_FOLDER'] = 'exports'
app.config['SECRET_KEY'] = 'obsid-selenium-automation-secret-2025'

# Configurare Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Model User simplu
class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)

# Inițializare SocketIO pentru WebSocket live logs
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')

# Queue-uri globale pentru comunicare între threads
automation_logs_queue = queue.Queue()
automation_input_queue = queue.Queue()
automation_active = False
stop_requested = False
current_automation_instance = None # Referință către instanța curentă de automatizare

# Creare directoare dacă nu există
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    """Verifică dacă fișierul are extensie permisă"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extrageInfoProdus(text_produs):
    """
    Extrage informații din textul produsului
    Returns: (nume_parfum, cantitate_ml, numar_bucati) sau None dacă nu e decant
    """
    if 'Decant' not in text_produs:
        return None

    # Extrage cantitatea în ml (3 ml, 5 ml, 10 ml)
    match_ml = re.search(r'Decant (\d+) ml parfum (.+?),', text_produs)
    if not match_ml:
        return None

    cantitate_ml = int(match_ml.group(1))
    nume_parfum = match_ml.group(2).strip()

    # Extrage numărul de bucăți
    match_bucati = re.search(r'(\d+\.\d+)$', text_produs.strip())
    if match_bucati:
        numar_bucati = float(match_bucati.group(1))
    else:
        numar_bucati = 1.0

    return (nume_parfum, cantitate_ml, int(numar_bucati))


def extrageInfoProdusIntreg(text_produs):
    """
    Extrage informații pentru produse întregi (non-decanturi)
    Returns: (nume_produs, numar_bucati) sau None
    """
    if 'Decant' in text_produs:
        return None
        
    # Extrage numărul de bucăți de la final (ex: ", 1.00")
    match_bucati = re.search(r'(\d+\.\d+)$', text_produs.strip())
    if match_bucati:
        numar_bucati = float(match_bucati.group(1))
        # Scoate cantitatea din nume
        nume_produs = text_produs.rsplit(',', 1)[0].strip()
    else:
        numar_bucati = 1.0
        nume_produs = text_produs.strip()
        
    return (nume_produs, int(numar_bucati))


def detecteazaColoane(df):
    """
    Detectează automat coloanele necesare din Excel
    Returns: (coloana_status, coloana_produse)
    """
    coloane = df.columns.tolist()

    # Detectare coloană Status
    coloana_status = None
    for col in coloane:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in ['status', 'stare', 'statu']):
            coloana_status = col
            break

    # Detectare coloană Produse
    coloana_produse = None
    for col in coloane:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in ['produse', 'produs', 'articol', 'item']):
            coloana_produse = col
            break

    if not coloana_status:
        raise ValueError('Nu s-a găsit coloana cu statusul comenzii (trebuie să conțină "Status" în nume)')

    if not coloana_produse:
        raise ValueError('Nu s-a găsit coloana cu produsele comandate (trebuie să conțină "Produse" în nume)')

    return coloana_status, coloana_produse


def proceseazaComenzi(fisier_path):
    """
    Procesează fișierul cu comenzi și returnează raportul de producție
    Detectează automat coloanele necesare (ordinea nu mai contează)
    Returnează și SKU-urile pentru fiecare produs
    """
    # Asigură-te că baza de date este încărcată
    if not PRODUCT_DB:
        logger.warning("⚠ Baza de date produse este goală! Încerc reîncărcarea...")
        load_product_db()

    df = pd.read_excel(fisier_path)

    # Detectare automată coloane
    coloana_status, coloana_produse = detecteazaColoane(df)

    # Verificare coloană atribute (pentru SKU)
    coloana_atribute = None
    for col in df.columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in ['atribute', 'atribut']):
            coloana_atribute = col
            break

    # Filtrare comenzi finalizate (flexibil - acceptă variante)
    df_finalizate = df[df[coloana_status].astype(str).str.contains('Finalizata|Confirmata', case=False, na=False)]

    # Raport agregat pe SKU (în loc de doar pe nume+cantitate)
    raport = defaultdict(lambda: {'nume': '', 'cantitate_ml': 0, 'bucati': 0})
    
    # Raport separat pentru produse întregi
    raport_intregi = defaultdict(lambda: {'nume': '', 'bucati': 0})

    # Log debug pentru DB
    logger.info(f"🔍 Procesare comenzi... DB size: {len(PRODUCT_DB)}")

    for idx, row in df_finalizate.iterrows():
        produse_text = str(row[coloana_produse])
        # atribute_text = str(row[coloana_atribute]) if coloana_atribute else '' # Nu mai folosim atribute

        produse = produse_text.split(' | ')

        for i, produs in enumerate(produse):
            # 1. Încearcă procesare ca DECANT
            info = extrageInfoProdus(produs.strip())
            if info:
                nume_parfum, cantitate_ml, numar_bucati = info
                
                # Căutare SKU
                found_sku = 'N/A'
                
                # 1. Căutare în baza de date produse (Prioritate MAXIMĂ)
                # Curățăm numele produsului din comandă (scoatem cantitatea de la final ", 1.00")
                produs_clean = re.sub(r', \d+\.\d+$', '', produs.strip())
                produs_norm = normalize_name(produs_clean)
                
                if produs_norm in PRODUCT_DB:
                    found_sku = PRODUCT_DB[produs_norm]
                else:
                    logger.warning(f"⚠ Produs negăsit în DB: {produs_clean} (norm: {produs_norm}). SKU setat la N/A.")
                    found_sku = 'N/A'
                
                sku = found_sku

                # FILTRARE SUPLIMENTARĂ: Exclude produsele care nu sunt decanturi (nu au extensie -3/-5/-10)
                if sku != 'N/A' and not re.search(r'-\d+$', sku):
                    logger.info(f"Produs exclus (SKU non-decant): {nume_parfum} | SKU: {sku}")
                    continue
                
                if sku == 'N/A':
                     logger.warning(f"Produs cu SKU N/A acceptat: {nume_parfum}")

                # Agregare pe SKU
                raport[sku]['nume'] = nume_parfum
                raport[sku]['cantitate_ml'] = cantitate_ml
                raport[sku]['bucati'] += numar_bucati
            
            # 2. Dacă nu e decant, verifică dacă e PRODUS ÎNTREG
            else:
                info_intreg = extrageInfoProdusIntreg(produs.strip())
                if info_intreg:
                    nume_produs, numar_bucati = info_intreg
                    
                    # Căutare SKU pentru produs întreg
                    produs_clean = re.sub(r', \d+\.\d+$', '', produs.strip())
                    produs_norm = normalize_name(produs_clean)
                    
                    sku = PRODUCT_DB.get(produs_norm, 'N/A')
                    
                    # Agregare
                    key = sku if sku != 'N/A' else nume_produs
                    raport_intregi[key]['nume'] = nume_produs
                    raport_intregi[key]['bucati'] += numar_bucati
                    raport_intregi[key]['sku'] = sku

    return raport, raport_intregi, len(df_finalizate), len(df)


def genereazaTabelRaport(raport):
    """
    Generează datele pentru tabel în format optimizat
    (parfumul apare o singură dată pe coloana A, cu SKU inclus)
    """
    # Grupare pe parfum
    parfumuri = {}
    for sku, info in raport.items():
        nume_parfum = info['nume']
        cantitate_ml = info['cantitate_ml']
        bucati = info['bucati']

        if nume_parfum not in parfumuri:
            parfumuri[nume_parfum] = {}
        if cantitate_ml not in parfumuri[nume_parfum]:
            parfumuri[nume_parfum][cantitate_ml] = []

        parfumuri[nume_parfum][cantitate_ml].append({
            'sku': sku,
            'bucati': bucati
        })

    # Construire rânduri pentru tabel
    randuri = []
    for nume_parfum in sorted(parfumuri.keys()):
        cantitati = parfumuri[nume_parfum]
        total_bucati = sum(sum(item['bucati'] for item in items) for items in cantitati.values())

        # Sortare cantități
        cantitati_sortate = sorted(cantitati.items())

        # Primul rând - cu numele parfumului
        primul_rand = True
        for ml, items in cantitati_sortate:
            for item in items:
                randuri.append({
                    'parfum': nume_parfum if primul_rand else '',
                    'sku': item['sku'],
                    'cantitate_ml': ml,
                    'bucati': item['bucati'],
                    'total': total_bucati if primul_rand else '',
                    'este_prim': primul_rand
                })
                primul_rand = False

    return randuri


def proceseazaBonuriProductie(fisier_path):
    """
    Procesează fișierul și extrage bonuri de producție agregate pe SKU
    Returns: lista de bonuri cu SKU, nume produs, cantitate agregată
    """
    df = pd.read_excel(fisier_path)

    # Detectare automată coloane
    coloana_status, coloana_produse = detecteazaColoane(df)

    # Verificare coloană atribute
    coloana_atribute = None
    for col in df.columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in ['atribute', 'atribut']):
            coloana_atribute = col
            break

    if not coloana_atribute:
        raise ValueError('Nu s-a găsit coloana cu atributele produselor')

    # Filtrare comenzi finalizate
    df_finalizate = df[df[coloana_status].astype(str).str.contains('Finalizata|Confirmata', case=False, na=False)]

    # Agregare bonuri pe SKU
    bonuri_agregate = defaultdict(lambda: {'nume': '', 'cantitate': 0, 'comenzi': []})

    for idx, row in df_finalizate.iterrows():
        produse_text = str(row[coloana_produse])
        atribute_text = str(row[coloana_atribute])

        # Split produse
        produse = produse_text.split(' | ')

        # Extrage SKU-uri din atribute
        sku_matches = re.findall(r'([^,\s]+):\s*\(', atribute_text)

        # Match produse cu SKU-uri
        for i, produs in enumerate(produse):
            produs = produs.strip()

            # Doar decanturi
            if 'Decant' not in produs:
                continue

            # Extrage cantitatea din produs
            match_cantitate = re.search(r'(\d+\.\d+)$', produs)
            cantitate = int(float(match_cantitate.group(1))) if match_cantitate else 1

            # SKU pentru acest produs
            sku = sku_matches[i] if i < len(sku_matches) else 'N/A'

            # Extrage numele parfumului
            match_parfum = re.search(r'Decant (\d+) ml parfum (.+?),', produs)
            if match_parfum:
                ml = match_parfum.group(1)
                nume_parfum = match_parfum.group(2)
                nume_complet = f"Decant {ml}ml {nume_parfum}"
            else:
                nume_complet = produs[:60]

            # Agregare
            bonuri_agregate[sku]['nume'] = nume_complet
            bonuri_agregate[sku]['cantitate'] += cantitate
            bonuri_agregate[sku]['comenzi'].append(str(row.get('Numar Comanda', '')))

    # Sortare după cantitate (descrescător)
    bonuri_sortate = sorted(bonuri_agregate.items(), key=lambda x: x[1]['cantitate'], reverse=True)

    # Convertire la format pentru JSON
    rezultat = []
    for sku, info in bonuri_sortate:
        comenzi_unice = list(set(info['comenzi']))[:5]
        rezultat.append({
            'sku': sku,
            'nume': info['nume'],
            'cantitate': info['cantitate'],
            'comenzi': comenzi_unice,
            'total_comenzi': len(set(info['comenzi']))
        })

    return rezultat


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Credențiale din ENV sau default
        valid_user = os.environ.get('APP_USER', 'admin')
        valid_pass = os.environ.get('APP_PASS', 'obsid123')
        
        # Debug logs
        logger.info(f"LOGIN DEBUG: EnvUser='{valid_user}', EnvPassLen={len(valid_pass) if valid_pass else 0}")
        
        if username and password and valid_user and valid_pass:
            # Comparare directă (fără strip pentru parolă dacă conține spații intenționate, dar strip la user e ok)
            if username.strip() == valid_user.strip() and password == valid_pass:
                user = User(1)
                login_user(user)
                next_page = request.args.get('next')
                return redirect(next_page or url_for('index'))
            else:
                # DEBUG EXTINS PENTRU UI - Arată ce așteaptă serverul vs ce a primit
                # ATENȚIE: A se scoate în producție după debug!
                debug_msg = (
                    f"DEBUG INFO: "
                    f"Expected User: '{valid_user}' | "
                    f"Expected Pass Len: {len(valid_pass)} | "
                    f"Received Pass Len: {len(password)} | "
                    f"Pass Match: {password == valid_pass}"
                )
                return render_template('login.html', error=f'Invalid credentials. {debug_msg}')
        else:
             return render_template('login.html', error='Missing credentials configuration')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """Pagina principală"""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """
    Upload și procesare fișier Excel
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Nu a fost selectat niciun fișier'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Nu a fost selectat niciun fișier'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Tipul fișierului nu este permis. Folosiți .xlsx sau .xls'}), 400

    try:
        # Salvare fișier temporar
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], save_filename)
        file.save(filepath)

        # Procesare comenzi (returnează raport cu SKU-uri)
        raport, raport_intregi, comenzi_finalizate, total_comenzi = proceseazaComenzi(filepath)

        # Generare tabel pentru Tab 1 (Raport Decanturi)
        randuri = genereazaTabelRaport(raport)

        # Calcul sumar pentru Tab 1
        sumar_cantitati = defaultdict(int)
        for sku, info in raport.items():
            sumar_cantitati[info['cantitate_ml']] += info['bucati']

        total_decanturi = sum(sumar_cantitati.values())

        # Generare date pentru Tab 2 (Bonuri de Producție)
        # Sortare după cantitate (descrescător)
        bonuri_sortate = sorted(raport.items(), key=lambda x: x[1]['bucati'], reverse=True)
        bonuri = []
        for sku, info in bonuri_sortate:
            bonuri.append({
                'sku': sku,
                'nume': f"Decant {info['cantitate_ml']}ml {info['nume']}",
                'cantitate': info['bucati']
            })

        # Pregătire răspuns (include date pentru ambele taburi)
        return jsonify({
            'success': True,
            'filename': save_filename,
            'comenzi_finalizate': comenzi_finalizate,
            'total_comenzi': total_comenzi,
            'comenzi_anulate': total_comenzi - comenzi_finalizate,
            # Date pentru Tab 1 (Raport Decanturi)
            'randuri': randuri,
            'sumar': {
                'cantitati': dict(sorted(sumar_cantitati.items())),
                'total': total_decanturi
            },
            # Date pentru Tab 2 (Bonuri de Producție)
            'bonuri': bonuri,
            'total_bonuri': len(bonuri),
            'total_bucati': total_decanturi,
            # Date pentru Produse Întregi (pentru export)
            'produse_intregi': [
                {'sku': info['sku'], 'nume': info['nume'], 'bucati': info['bucati']}
                for info in sorted(raport_intregi.values(), key=lambda x: x['nume'])
            ]
        })

    except Exception as e:
        return jsonify({'error': f'Eroare la procesare: {str(e)}'}), 500


@app.route('/process-vouchers', methods=['POST'])
@login_required
def process_vouchers():
    """
    Procesare fișier pentru bonuri de producție (agregate pe SKU)
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Nu a fost selectat niciun fișier'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Nu a fost selectat niciun fișier'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Tipul fișierului nu este permis. Folosiți .xlsx sau .xls'}), 400

    try:
        # Salvare fișier temporar
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], save_filename)
        file.save(filepath)

        # Procesare bonuri de producție
        bonuri = proceseazaBonuriProductie(filepath)

        total_bonuri = len(bonuri)
        total_bucati = sum(bon['cantitate'] for bon in bonuri)

        # Pregătire răspuns
        return jsonify({
            'success': True,
            'filename': save_filename,
            'bonuri': bonuri,
            'total_bonuri': total_bonuri,
            'total_bucati': total_bucati
        })

    except Exception as e:
        return jsonify({'error': f'Eroare la procesare: {str(e)}'}), 500


@app.route('/export/<filename>')
@login_required
def export_excel(filename):
    """
    Export raport în Excel
    """
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        if not os.path.exists(filepath):
            return jsonify({'error': 'Fișierul nu există'}), 404

        # Procesare comenzi (returnează raport cu SKU-uri)
        raport, raport_intregi, comenzi_finalizate, total_comenzi = proceseazaComenzi(filepath)

        # Pregătire date pentru Excel (include SKU)
        date_raport = []
        for sku, info in sorted(raport.items(), key=lambda x: (x[1]['nume'], x[1]['cantitate_ml'])):
            date_raport.append({
                'SKU': sku,
                'Parfum': info['nume'],
                'Cantitate (ml)': info['cantitate_ml'],
                'Bucăți': info['bucati']
            })

        df_raport = pd.DataFrame(date_raport)

        # Pregătire date pentru Produse Întregi
        date_intregi = []
        for key, info in sorted(raport_intregi.items(), key=lambda x: x[1]['nume']):
            date_intregi.append({
                'SKU': info['sku'],
                'Parfum': info['nume'],
                'Cantitate (ml)': 'FULL SIZE',
                'Bucăți': info['bucati']
            })
        
        df_intregi = pd.DataFrame(date_intregi)

        if df_raport.empty:
            # DataFrame gol - structură default pentru a evita erori
            df_raport = pd.DataFrame(columns=['SKU', 'Parfum', 'Cantitate (ml)', 'Bucăți'])
            df_sumar = pd.DataFrame(columns=['Parfum', 'Total Bucăți'])
        else:
            # Sumar pe parfum
            df_sumar = df_raport.groupby('Parfum').agg({'Bucăți': 'sum'}).reset_index()
            df_sumar.columns = ['Parfum', 'Total Bucăți']

        # Creare Excel în memorie
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Scrie Raportul Detaliat (Decanturi)
            df_raport.to_excel(writer, sheet_name='Raport Detaliat', index=False)
            
            # 2. Dacă avem produse întregi, le adăugăm după 5 rânduri goale
            if not df_intregi.empty:
                worksheet = writer.sheets['Raport Detaliat']
                start_row = len(df_raport) + 7 # +1 header + 1 index-0 + 5 empty rows
                
                # Scrie Header pentru secțiunea nouă
                worksheet.cell(row=start_row-1, column=1, value="PRODUSE ÎNTREGI (FULL SIZE)")
                
                # Scrie datele (folosind to_excel cu startrow)
                # ATENȚIE: Pandas to_excel suprascrie sheet-ul dacă nu folosim mode='a' și if_sheet_exists='overlay'
                # Dar openpyxl engine permite scrierea directă.
                # Totuși, cea mai simplă metodă cu Pandas este să scriem totul deodată sau să folosim openpyxl direct.
                
                # Metoda sigură: Scriem manual cu openpyxl peste sheet-ul existent
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Header tabel
                header_row = ['SKU', 'Parfum', 'Cantitate (ml)', 'Bucăți']
                for col_idx, value in enumerate(header_row, 1):
                    worksheet.cell(row=start_row, column=col_idx, value=value)
                
                # Date tabel
                for r_idx, row in enumerate(dataframe_to_rows(df_intregi, index=False, header=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        worksheet.cell(row=start_row + r_idx, column=c_idx, value=value)

            df_sumar.to_excel(writer, sheet_name='Sumar pe Parfum', index=False)

        output.seek(0)

        # Nume fișier export
        export_filename = f"raport_productie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=export_filename
        )

    except Exception as e:
        return jsonify({'error': f'Eroare la export: {str(e)}'}), 500


@app.route('/download-model')
@login_required
def download_model():
    """
    Download fișier model de import
    """
    try:
        # Creare model în memorie
        data = {
            'Status Comanda': [
                'Comanda Finalizata (Facturata)',
                'Comanda Finalizata (Facturata)',
                'Anulata'
            ],
            'Produse comandate': [
                'Decant 3 ml parfum Yara Lattafa, parfum femei, 1.00 | Decant 5 ml parfum Eclaire Lattafa, parfum femei, 2.00',
                'Decant 10 ml parfum Yum Yum Armaf, parfum femei, 1.00',
                'Decant 3 ml parfum Khamrah Lattafa, unisex, 1.00'
            ]
        }

        df_model = pd.DataFrame(data)

        # Creare Excel în memorie
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_model.to_excel(writer, sheet_name='Model Import', index=False)

            # Adaugă sheet cu instrucțiuni
            instructiuni = pd.DataFrame({
                'INSTRUCȚIUNI UTILIZARE': [
                    '1. Coloane OBLIGATORII:',
                    '   - Status Comanda: trebuie să conțină "Finalizata" pentru comenzi procesate',
                    '   - Produse comandate: lista de produse separate prin " | "',
                    '',
                    '2. Coloanele pot fi în orice ordine',
                    '',
                    '3. Format produse:',
                    '   Decant [CANTITATE] ml parfum [NUME PARFUM], [sex], [NUMĂR BUCĂȚI]',
                    '',
                    '4. Exemple valide:',
                    '   - Decant 3 ml parfum Yara Lattafa, parfum femei, 1.00',
                    '   - Decant 10 ml parfum Khamrah Lattafa, unisex, 2.00',
                    '',
                    '5. Multiple produse separate prin " | "',
                    '',
                    '6. Comenzile anulate sunt automat excluse din raport'
                ]
            })
            instructiuni.to_excel(writer, sheet_name='Instrucțiuni', index=False)

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='model_import_obsid_decanturi.xlsx'
        )

    except Exception as e:
        return jsonify({'error': f'Eroare la generare model: {str(e)}'}), 500


@app.route('/health')
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'service': 'OBSID Decant Manager'})


@app.route('/start-automation-selenium', methods=['POST'])
@login_required
def start_automation_selenium():
    """
    Pornește automatizarea Selenium pentru crearea bonurilor în Oblio
    """
    try:
        data = request.get_json()
        bonuri = data.get('bonuri', [])
        oblio_cookies = data.get('oblio_cookies')  # Cookies trimise din frontend

        if not bonuri:
            return jsonify({'error': 'Nu există bonuri de procesat'}), 400

        # Import automation class
        from automatizare_oblio_selenium import OblioAutomation
        import platform

        # Detectează dacă rulează în container/Linux (server) sau Windows (local)
        is_linux = platform.system() == 'Linux'

        # Inițializare automation cu parametri corecți
        automation = OblioAutomation(
            use_existing_profile=not is_linux,  # False pe server, True pe Windows local
            headless=is_linux  # True pe server (headless), False pe Windows (cu GUI)
        )

        # Setup driver
        if not automation.setup_driver():
            error_msg = 'Nu s-a putut porni Chrome WebDriver.'
            if is_linux:
                error_msg += ' Verifică logs în automatizare_oblio.log pentru detalii.'
            else:
                error_msg += ' Verifică că Chrome este instalat și rulează cu --remote-debugging-port=9222'

            return jsonify({
                'error': error_msg,
                'hint': 'Windows: Pornește Chrome cu remote debugging. Linux: Verifică logs.'
            }), 500

        # Citește credențialele Oblio din environment variables
        oblio_email = os.environ.get('OBLIO_EMAIL')
        oblio_password = os.environ.get('OBLIO_PASSWORD')
        
        # Log informații despre autentificare
        if oblio_cookies and len(oblio_cookies) > 0:
            logger.info(f"🍪 Cookies primite din frontend: {len(oblio_cookies)} cookies")
            logger.info("🍪 Autentificare cu cookies din frontend")
        else:
            logger.info("⚠️ Cookies Oblio NU au fost primite (normal - sunt HttpOnly)")
            
        if oblio_email and oblio_password:
            logger.info("🔐 Credențiale Oblio disponibile din environment variables")
            logger.info("🔐 Voi folosi autentificare cu email/password")
        else:
            # Verifică că avem cel puțin credențiale
            if not (oblio_email and oblio_password):
                return jsonify({
                    'error': 'Credențiale Oblio lipsă! Setează OBLIO_EMAIL și OBLIO_PASSWORD în environment variables.',
                    'hint': 'Cookies HttpOnly nu pot fi accesate din JavaScript - folosim email/password.'
                }), 500
        
        # Procesează bonurile
        stats = automation.process_bonuri(bonuri, oblio_cookies, oblio_email, oblio_password)

        # Închide browser
        automation.close()

        # Returnează rezultatul
        return jsonify({
            'success': True,
            'stats': stats,
            'message': f'Automatizare finalizată! {stats["success"]}/{stats["total"]} bonuri create cu succes.'
        })

    except Exception as e:
        return jsonify({'error': f'Eroare la automatizare: {str(e)}'}), 500


# ============================================================
# WEBSOCKET HANDLERS - Live Terminal cu Interactive Input
# ============================================================

@socketio.on('connect')
def handle_connect():
    """Client conectat la WebSocket"""
    logger.info(f"🔌 Client conectat la WebSocket")
    emit('log', {'type': 'info', 'message': '✅ Conectat la terminal live!'})


@socketio.on('disconnect')
def handle_disconnect():
    """Client deconectat de la WebSocket"""
    logger.info(f"🔌 Client deconectat de la WebSocket")


@socketio.on('stop_automation')
def handle_stop_automation():
    global stop_requested, automation_active
    if automation_active:
        stop_requested = True
        logger.info("🛑 Cerere de oprire automatizare primită")
        emit('log', {'type': 'warning', 'message': '⚠️ Se încearcă oprirea automatizării...'})
    else:
        emit('log', {'type': 'info', 'message': 'ℹ️ Nu există nicio automatizare activă.'})
        # Force reset just in case
        automation_active = False


@socketio.on('start_automation_live')
def handle_start_automation_live(data):
    """
    Pornește automatizarea cu logs live și interactive input
    """
    global automation_active, stop_requested
    stop_requested = False

    if automation_active:
        emit('log', {'type': 'error', 'message': '⚠️ O automatizare este deja în desfășurare!'})
        emit('automation_status', {'active': True})
        return

    bonuri = data.get('bonuri', [])
    force_mode = data.get('force_mode', False)

    if not bonuri:
        emit('log', {'type': 'error', 'message': '❌ Nu există bonuri de procesat'})
        return

    automation_active = True

    if force_mode:
        emit('log', {'type': 'warning', 'message': f'🚀 START AUTOMATIZARE: {len(bonuri)} bonuri (FORCE MODE: Processing ALL vouchers)'})
    else:
        emit('log', {'type': 'info', 'message': f'🚀 START AUTOMATIZARE: {len(bonuri)} bonuri'})

    # DEBUG: Test emit înainte de thread
    emit('log', {'type': 'warning', 'message': '⚡ IMEDIAT PORNESC THREAD-UL...'})

    # Pornește automation cu socketio.start_background_task (FUNCȚIONEAZĂ cu eventlet!)
    try:
        socketio.start_background_task(
            run_automation_with_live_logs,
            bonuri,
            request.sid,
            force_mode
        )
        emit('log', {'type': 'success', 'message': '✅ BACKGROUND TASK PORNIT! Așteaptă logs...'})
    except Exception as e:
        emit('log', {'type': 'error', 'message': f'❌ EROARE LA PORNIRE TASK: {str(e)}'})
        automation_active = False


@socketio.on('user_input')
def handle_user_input(data):
    """
    Primește input de la utilizator (email, password, 2FA code)
    """
    input_type = data.get('type')  # 'email', 'password', '2fa'
    value = data.get('value')

    logger.info(f"📥 Primit input de la user: type={input_type}")

    # Pune input-ul în queue pentru ca Selenium să-l preia
    automation_input_queue.put({
        'type': input_type,
        'value': value
    })

    emit('log', {'type': 'success', 'message': f'✅ Input primit: {input_type}'})


def send_heartbeat(client_sid, stop_event):
    """
    Trimite heartbeat la fiecare 5 secunde pentru a menține conexiunea vie
    """
    while not stop_event.is_set():
        try:
            socketio.emit('heartbeat', {'timestamp': time.time()}, room=client_sid)
            eventlet.sleep(5)  # Heartbeat la fiecare 5 secunde (mai frecvent!)
        except:
            break


def run_automation_with_live_logs(bonuri, client_sid, force_mode=False):
    """
    Rulează automatizarea în background și trimite logs live
    FOLOSEȘTE socketio.start_background_task() care funcționează cu eventlet!
    Args:
        bonuri: Lista de bonuri de procesat
        client_sid: Session ID al clientului WebSocket
        force_mode: Dacă True, procesează TOATE bonurile fără verificare duplicate
    """
    global automation_active, stop_requested

    # Event pentru oprirea heartbeat
    import threading
    stop_heartbeat = threading.Event()

    # Pornește heartbeat în background
    socketio.start_background_task(send_heartbeat, client_sid, stop_heartbeat)

    # Cu socketio.start_background_task() NU mai trebuie app.app_context()!
    try:
        socketio.emit('log', {
            'type': 'info',
            'message': '🔧 START BACKGROUND TASK - Pornire automation...'
        }, room=client_sid)

        # CRITICAL: Try to import with detailed error
        try:
            socketio.emit('log', {
                'type': 'info',
                'message': '📥 Încerc să importez automatizare_oblio_selenium...'
            }, room=client_sid)

            from automatizare_oblio_selenium import OblioAutomation

            socketio.emit('log', {
                'type': 'success',
                'message': '✅ OblioAutomation importat cu succes!'
            }, room=client_sid)
        except ImportError as ie:
            socketio.emit('log', {
                'type': 'error',
                'message': f'❌ IMPORT ERROR: {str(ie)}'
            }, room=client_sid)
            raise
        except Exception as ie:
            socketio.emit('log', {
                'type': 'error',
                'message': f'❌ EROARE LA IMPORT: {str(ie)}'
            }, room=client_sid)
            raise

        import platform

        socketio.emit('log', {
            'type': 'info',
            'message': '📦 Import-uri OK - Inițializare Selenium...'
        }, room=client_sid)

        is_linux = platform.system() == 'Linux'

        socketio.emit('log', {
            'type': 'info',
            'message': f'🖥️ Sistem detectat: {"Linux (Headless)" if is_linux else "Windows (Visual)"}'
        }, room=client_sid)

        # Inițializare automation cu logs live
        automation = OblioAutomation(
            use_existing_profile=not is_linux,
            headless=is_linux,
            log_callback=lambda msg, level: socketio.emit('log', {
                'type': level,
                'message': msg,
                'timestamp': time.time()
            }, room=client_sid),
            input_callback=lambda prompt: wait_for_user_input(prompt, client_sid)
        )
        
        # Setăm instanța globală pentru a putea fi oprită
        global current_automation_instance
        current_automation_instance = automation

        # Setup driver
        if not automation.setup_driver():
            socketio.emit('log', {
                'type': 'error',
                'message': '❌ Nu s-a putut porni Chrome WebDriver'
            }, room=client_sid)
            socketio.emit('automation_complete', {
                'success': False,
                'error': 'ChromeDriver failed to start'
            }, room=client_sid)
            return

        # Procesează bonuri BON CU BON (fix timeout!)
        stats = {
            'total': len(bonuri),
            'success': 0,
            'failed': 0,
            'errors': []
        }

        # Credențiale Oblio
        oblio_email = os.environ.get('OBLIO_EMAIL')
        oblio_password = os.environ.get('OBLIO_PASSWORD')
        
        if oblio_email:
            socketio.emit('log', {
                'type': 'info',
                'message': f'🔐 Folosesc credențiale din ENV: {oblio_email}'
            }, room=client_sid)

        # --- SMART RESUME: Verifică ce s-a lucrat deja azi ---
        if force_mode:
            socketio.emit('log', {
                'type': 'warning',
                'message': '⚠️ FORCE MODE ACTIVAT: Procesez TOATE bonurile fără verificare duplicate!'
            }, room=client_sid)

        if not force_mode:
            try:
                # 1. Verificare în Baza de Date (PostgreSQL) - Prioritar
                processed_db = database.get_bonuri_azi()
                processed_skus_db = {item['sku'] for item in processed_db}

                if processed_skus_db:
                    socketio.emit('log', {
                        'type': 'info',
                        'message': f'📊 Găsite {len(processed_skus_db)} bonuri în baza de date locală.'
                    }, room=client_sid)

                # 2. Verificare în Oblio (Scraping) - Fallback / Validare
                # Facem asta doar dacă DB-ul e gol sau pentru siguranță maximă
                processed_texts_oblio = []
                if automation.login_if_needed(oblio_email, oblio_password):
                    processed_texts_oblio = automation.get_todays_processed_texts()

                # Filtrare
                initial_count = len(bonuri)
                bonuri_filtrate = []

                for bon in bonuri:
                    sku = bon.get('sku', '')
                    nume = bon.get('nume', '')

                    is_processed = False

                    # A. Verificare DB
                    if sku in processed_skus_db:
                        is_processed = True
                        logger.info(f"⏭️ Skip (DB): {nume} (SKU: {sku})")

                    # B. Verificare Oblio (dacă nu e găsit în DB)
                    if not is_processed and processed_texts_oblio:
                        for text in processed_texts_oblio:
                            if sku and len(sku) > 3 and sku in text:
                                is_processed = True
                                break
                            if nume and len(nume) > 5:
                                # Match mai relaxat pe nume
                                match_parfum = re.search(r'Decant \d+ ?ml (parfum )?(.+)', nume, re.IGNORECASE)
                                if match_parfum:
                                    nume_parfum_doar = match_parfum.group(2).strip()
                                    if len(nume_parfum_doar) > 4 and nume_parfum_doar in text:
                                        match_ml = re.search(r'Decant (\d+)', nume)
                                        if match_ml and match_ml.group(1) in text:
                                            is_processed = True
                                            break

                        if is_processed:
                            logger.info(f"⏭️ Skip (Oblio): {nume} (SKU: {sku})")
                            # Opțional: Salvăm în DB dacă am găsit în Oblio dar nu era în DB
                            try:
                                database.adauga_bon(sku, nume, bon.get('cantitate', 1))
                            except: pass

                    if not is_processed:
                        bonuri_filtrate.append(bon)

                bonuri = bonuri_filtrate
                skipped_count = initial_count - len(bonuri)

                if skipped_count > 0:
                    socketio.emit('log', {
                        'type': 'warning',
                        'message': f'⏭️ SMART RESUME: Am sărit peste {skipped_count} bonuri deja create astăzi.'
                    }, room=client_sid)

                    stats['total'] = len(bonuri)
                    stats['skipped'] = skipped_count
                
                if len(bonuri) == 0:
                    socketio.emit('log', {
                        'type': 'success',
                        'message': '✅ Toate bonurile din listă au fost deja procesate astăzi!'
                    }, room=client_sid)
                
        except Exception as e:
            logger.error(f"Eroare Smart Resume: {e}")
            socketio.emit('log', {
                'type': 'warning',
                'message': f'⚠️ Eroare la verificarea istoricului: {e}'
            }, room=client_sid)
        # ----------------------------------------------------

        # Procesare BON cu BON cu progress live (BATCH OPTIMIZATION)
        batch_size = 5 # Procesăm câte 5 bonuri în paralel (crescut de la 3)
        retryable_bonuri = [] # Lista pentru bonuri care pot fi reîncercate (timeout, erori rețea)
        
        for i in range(0, len(bonuri), batch_size):
            if stop_requested:
                break
                
            batch = bonuri[i:i + batch_size]
            batch_indices = range(i + 1, i + len(batch) + 1)
            
            socketio.emit('log', {
                'type': 'info',
                'message': f'🚀 Procesare batch {i//batch_size + 1}: {len(batch)} bonuri în paralel...'
            }, room=client_sid)
            
            # Emit progress pentru toate din batch
            for idx, bon in zip(batch_indices, batch):
                socketio.emit('progress', {
                    'current': idx,
                    'total': len(bonuri),
                    'sku': bon.get('sku'),
                    'nume': bon.get('nume', ''),
                    'cantitate': bon.get('cantitate', 1)
                }, room=client_sid)
            
            eventlet.sleep(0.1)

            try:
                # Apelăm metoda de batch
                results = automation.create_production_vouchers_batch(
                    batch,
                    None,   # cookies
                    oblio_email,
                    oblio_password
                )
                
                # Procesăm rezultatele
                for res in results:
                    sku = res['sku']
                    success = res['success']
                    msg = res['message']
                    
                    # Găsim indexul original
                    original_idx = next((idx for idx, b in zip(batch_indices, batch) if b.get('sku') == sku), 0)
                    original_bon = next((b for b in batch if b.get('sku') == sku), {})
                    
                    if success:
                        stats['success'] += 1
                        stats.setdefault('successful_products', []).append(original_bon)
                        
                        socketio.emit('bon_complete', {
                            'index': original_idx,
                            'total': len(bonuri),
                            'success': True,
                            'sku': sku,
                            'message': f'✅ Bon {original_idx}/{len(bonuri)} finalizat cu succes!'
                        }, room=client_sid)
                    else:
                        stats['failed'] += 1
                        stats['errors'].append({'sku': sku, 'error': msg})
                        
                        # Verificăm dacă eroarea este retryable (NU este stoc insuficient)
                        if "stoc insuficient" not in msg.lower():
                            retryable_bonuri.append(original_bon)
                            logger.info(f"🔄 Bon adăugat la coada de retry: {sku} (Eroare: {msg})")
                        
                        socketio.emit('bon_complete', {
                            'index': original_idx,
                            'total': len(bonuri),
                            'success': False,
                            'sku': sku,
                            'message': f'❌ Bon {original_idx}/{len(bonuri)} eșuat: {msg}'
                        }, room=client_sid)
                        
                eventlet.sleep(0.5)

            except Exception as e:
                logger.error(f"❌ Eroare la procesarea batch-ului: {e}")
                # Mark all in batch as failed
                for idx, bon in zip(batch_indices, batch):
                    stats['failed'] += 1
                    stats['errors'].append({'sku': bon.get('sku'), 'error': str(e)})
                    
                    # Adăugăm tot batch-ul la retry
                    retryable_bonuri.append(bon)
                    
                    socketio.emit('bon_complete', {
                        'index': idx,
                        'total': len(bonuri),
                        'success': False,
                        'sku': bon.get('sku'),
                        'message': f'❌ Eroare batch: {str(e)[:100]}'
                    }, room=client_sid)

        # ============================================================
        # ETAPA 1.5: RETRY PENTRU BONURI EȘUATE (Timeout, Erori rețea)
        # ============================================================
        if retryable_bonuri and not stop_requested:
            socketio.emit('log', {
                'type': 'warning',
                'message': f'🔄 START RETRY pentru {len(retryable_bonuri)} bonuri eșuate (fără probleme de stoc)...'
            }, room=client_sid)
            
            # Procesăm retry-urile secvențial sau în batch-uri mici (2) pentru siguranță
            retry_batch_size = 2
            
            for i in range(0, len(retryable_bonuri), retry_batch_size):
                if stop_requested:
                    break
                    
                batch = retryable_bonuri[i:i + retry_batch_size]
                
                socketio.emit('log', {
                    'type': 'info',
                    'message': f'🔄 Retry batch {i//retry_batch_size + 1}: {len(batch)} bonuri...'
                }, room=client_sid)
                
                try:
                    results = automation.create_production_vouchers_batch(
                        batch,
                        None,
                        oblio_email,
                        oblio_password
                    )
                    
                    for res in results:
                        sku = res['sku']
                        success = res['success']
                        msg = res['message']
                        
                        original_bon = next((b for b in batch if b.get('sku') == sku), {})
                        
                        if success:
                            # Actualizăm stats: scădem din failed, adăugăm la success
                            stats['failed'] -= 1
                            stats['success'] += 1
                            stats.setdefault('successful_products', []).append(original_bon)
                            
                            # Scoatem eroarea veche din listă
                            stats['errors'] = [err for err in stats['errors'] if err['sku'] != sku]
                            
                            socketio.emit('log', {
                                'type': 'success',
                                'message': f'✅ RETRY REUȘIT pentru {sku}!'
                            }, room=client_sid)
                        else:
                            socketio.emit('log', {
                                'type': 'error',
                                'message': f'❌ RETRY EȘUAT pentru {sku}: {msg}'
                            }, room=client_sid)
                            
                except Exception as e:
                    logger.error(f"❌ Eroare la retry batch: {e}")

        # Check stop request outside loop
        if stop_requested:
            socketio.emit('log', {
                'type': 'warning',
                'message': '🛑 Automatizare oprită manual de utilizator!'
            }, room=client_sid)
            socketio.emit('automation_complete', {
                'success': False,
                'error': 'Stopped by user'
            }, room=client_sid)
            return # Exit function

        # ============================================================
        # ETAPA 2: TRANSFER GESTIUNE (Materiale consumabile -> Marfuri)
        # ============================================================
        successful_products = stats.get('successful_products', [])
        
        # FILTRARE: Păstrăm DOAR decanturile pentru transfer
        # Parfumurile întregi nu se transferă din consumabile
        products_to_transfer = []
        for prod in successful_products:
            sku = prod.get('sku', '')
            # Verificăm dacă SKU se termină în -3, -5, -10 (decanturi)
            is_decant_sku = bool(re.search(r'-(3|5|10)$', sku))
            
            # Fallback pe nume doar dacă nu avem SKU valid
            is_decant_name = 'Decant' in prod.get('nume', '')

            if is_decant_sku or (sku == 'N/A' and is_decant_name):
                products_to_transfer.append(prod)
            else:
                logger.info(f"Skip transfer produs non-decant: {prod.get('nume')} (SKU: {sku})")
                socketio.emit('log', {
                    'type': 'info',
                    'message': f'ℹ️ Skip transfer pentru produs întreg: {prod.get("nume")}'
                }, room=client_sid)

        if products_to_transfer and not stop_requested:
            socketio.emit('log', {
                'type': 'info',
                'message': f'🚚 START TRANSFER GESTIUNE pentru {len(products_to_transfer)} produse (din total {len(successful_products)})...'
            }, room=client_sid)
            eventlet.sleep(1)

            transfer_success = automation.create_transfer_note(products_to_transfer)
            
            if transfer_success:
                socketio.emit('log', {
                    'type': 'success',
                    'message': '✅ NOTA DE TRANSFER EMISĂ CU SUCCES!'
                }, room=client_sid)
            else:
                socketio.emit('log', {
                    'type': 'error',
                    'message': '❌ EROARE LA EMITEREA NOTEI DE TRANSFER!'
                }, room=client_sid)
        else:
            if not products_to_transfer:
                socketio.emit('log', {
                    'type': 'warning',
                    'message': '⚠️ Nu există decanturi de transferat (doar produse întregi sau lista e goală).'
                }, room=client_sid)

        # Închide browser
        automation.close()

        # Trimite rezultat final
        socketio.emit('automation_complete', {
            'success': True,
            'stats': stats,
            'message': f'✅ Automatizare finalizată! {stats["success"]}/{stats["total"]} bonuri create'
        }, room=client_sid)

    except Exception as e:
        logger.error(f"❌ Eroare în automation: {e}", exc_info=True)
        socketio.emit('log', {
            'type': 'error',
            'message': f'❌ EROARE: {str(e)}'
        }, room=client_sid)
        socketio.emit('automation_complete', {
            'success': False,
            'error': str(e)
        }, room=client_sid)
    finally:
        # Oprește heartbeat
        stop_heartbeat.set()
        automation_active = False


def wait_for_user_input(prompt, client_sid):
    """
    Așteaptă input de la utilizator prin WebSocket

    Args:
        prompt (dict): {'type': 'email'/'password'/'2fa', 'message': 'Enter email...'}
        client_sid: Socket ID pentru client

    Returns:
        str: Input-ul utilizatorului
    """
    # DEBUG logging
    logger.info(f"🔍 DEBUG wait_for_user_input: prompt={prompt}, client_sid={client_sid}")

    # Trimite prompt către frontend
    logger.info(f"📤 Emit 'input_required' către client {client_sid}")
    socketio.emit('input_required', prompt, room=client_sid)
    logger.info(f"✅ Emit trimis! Aștept răspuns în queue...")

    # Așteaptă răspuns în queue (cu timeout)
    try:
        user_input = automation_input_queue.get(timeout=300)  # 5 minute timeout
        logger.info(f"✅ Input primit din queue: {user_input}")
        return user_input.get('value')
    except queue.Empty:
        logger.error("⏱️ TIMEOUT - nu s-a primit input în 5 minute!")
        socketio.emit('log', {
            'type': 'error',
            'message': '⏱️ Timeout - nu s-a primit input de la utilizator'
        }, room=client_sid)
        return None


# ============================================================
# START APPLICATION
# ============================================================

if __name__ == '__main__':
    # Folosește socketio.run() în loc de app.run() pentru WebSocket support
    socketio.run(app, host='0.0.0.0', port=5000, debug=False)
